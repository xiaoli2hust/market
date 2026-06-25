"""营销活动查询路由。

提供活动列表、统计聚合、行为类型字典与 Excel 导出等能力，
是日报数据下沉到前端可视化的核心读路径。
"""

from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, distinct, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..auth import require_permission
from ..database import get_db
from ..models import Activity, Staff

router = APIRouter(prefix="/activities", tags=["activities"])


# ---------------------------------------------------------------------------
# 行为类型字典
# ---------------------------------------------------------------------------

ACTION_TYPES: list[dict[str, str]] = [
    {"value": "拜访客户", "label": "拜访客户", "color": "#1890ff"},
    {"value": "商机跟进", "label": "商机跟进", "color": "#fa8c16"},
    {"value": "方案撰写", "label": "方案撰写", "color": "#13c2c2"},
    {"value": "项目推进", "label": "项目推进", "color": "#52c41a"},
    {"value": "渠道拓展", "label": "渠道拓展", "color": "#722ed1"},
    {"value": "回款跟进", "label": "回款跟进", "color": "#f5222d"},
    {"value": "内部协作", "label": "内部协作", "color": "#8c8c8c"},
    {"value": "技术交流", "label": "技术交流", "color": "#2f54eb"},
    {"value": "POC测试", "label": "POC测试", "color": "#eb2f96"},
    {"value": "招投标", "label": "招投标", "color": "#faad14"},
    {"value": "合同谈判", "label": "合同谈判", "color": "#a0d911"},
    {"value": "客户维护", "label": "客户维护", "color": "#36cfc9"},
    {"value": "其他", "label": "其他", "color": "#bfbfbf"},
]


def _parse_date(value: str | None, field: str) -> date | None:
    """容错的日期字符串解析。"""

    if value is None or value == "":
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:  # pragma: no cover - 让 FastAPI 直接返回 400
        raise HTTPException(
            status_code=400,
            detail=f"invalid date for {field}: {value!r}, expected YYYY-MM-DD",
        ) from exc


def _activity_to_dict(activity: Activity, staff: Staff | None) -> dict[str, Any]:
    """将 ORM 对象拍平为前端友好的 dict（含 staff_name/department）。

    同时包含前端兼容字段别名（user_id/action_type/summary 等）。
    """

    report_date_str = activity.report_date.isoformat() if activity.report_date else None
    return {
        "id": activity.id,
        "staff_id": activity.staff_id,
        "user_id": activity.staff_id,  # 前端兼容别名
        "staff_name": staff.name if staff else None,
        "user_name": staff.name if staff else None,  # 前端兼容别名
        "department": staff.department if staff else None,
        "user_department": staff.department if staff else None,  # 前端兼容别名
        "report_date": report_date_str,
        "activity_date": report_date_str,  # 前端兼容别名
        "activity_type": activity.activity_type,
        "action_type": activity.activity_type,  # 前端兼容别名
        "action_type_label": activity.activity_type,  # 前端兼容别名（中文标签同值）
        "target": activity.target,
        "customer_name": activity.target,  # 前端兼容别名
        "opportunity": activity.opportunity,
        "opportunity_name": activity.opportunity,  # 前端兼容别名
        "opportunity_id": activity.opportunity_id,
        "description": activity.description,
        "summary": activity.description,  # 前端兼容别名
        "detail": activity.description,  # 前端兼容别名
        "confidence": activity.confidence,
        "is_reviewed": activity.is_reviewed,
        "source_file_id": activity.source_file_id,
        "source": "daily_report",  # 前端兼容
        "created_at": activity.created_at.isoformat() if activity.created_at else None,
    }


# ---------------------------------------------------------------------------
# 列表查询
# ---------------------------------------------------------------------------


@router.get("")
@router.get("/")
async def list_activities(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    staff_id: int | None = None,
    user_id: int | None = None,  # 前端兼容别名
    department: str | None = None,
    activity_type: str | None = None,
    action_type: str | None = None,  # 前端兼容别名
    opportunity: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    keyword: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(require_permission("dashboard:view")),
) -> dict[str, Any]:
    """多维度筛选活动记录，支持分页。

    返回 ``{"total": int, "items": [ActivityOut...]}``，按 ``report_date`` 倒序。
    """

    # 参数兼容：前端传 user_id/action_type，后端用 staff_id/activity_type
    effective_staff_id = staff_id or user_id
    effective_activity_type = activity_type or action_type

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date")

    conditions = []
    if effective_staff_id is not None:
        conditions.append(Activity.staff_id == effective_staff_id)
    if department:
        conditions.append(Staff.department == department)
    if effective_activity_type:
        conditions.append(Activity.activity_type == effective_activity_type)
    if opportunity:
        conditions.append(Activity.opportunity.ilike(f"%{opportunity}%"))
    if start is not None:
        conditions.append(Activity.report_date >= start)
    if end is not None:
        conditions.append(Activity.report_date <= end)
    if keyword:
        like = f"%{keyword}%"
        conditions.append(
            or_(
                Activity.description.ilike(like),
                Activity.target.ilike(like),
                Activity.opportunity.ilike(like),
            )
        )

    where_clause = and_(*conditions) if conditions else None

    # 总数
    count_stmt = select(func.count(Activity.id)).select_from(Activity).join(
        Staff, Staff.id == Activity.staff_id, isouter=True
    )
    if where_clause is not None:
        count_stmt = count_stmt.where(where_clause)
    total = (await db.execute(count_stmt)).scalar_one()

    # 列表
    list_stmt = (
        select(Activity, Staff)
        .join(Staff, Staff.id == Activity.staff_id, isouter=True)
    )
    if where_clause is not None:
        list_stmt = list_stmt.where(where_clause)
    list_stmt = (
        list_stmt.order_by(Activity.report_date.desc(), Activity.id.desc())
        .offset(skip)
        .limit(limit)
    )
    rows = (await db.execute(list_stmt)).all()
    items = [_activity_to_dict(activity, staff) for activity, staff in rows]

    return {"total": int(total or 0), "items": items, "list": items, "skip": skip, "limit": limit}


# ---------------------------------------------------------------------------
# 统计聚合
# ---------------------------------------------------------------------------


@router.get("/stats")
async def get_stats(
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(require_permission("dashboard:view")),
) -> dict[str, Any]:
    """活动统计聚合：总览 + 维度分布 + 近 7 天趋势。"""

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date")
    today = date.today()

    range_conditions: list = []
    if start is not None:
        range_conditions.append(Activity.report_date >= start)
    if end is not None:
        range_conditions.append(Activity.report_date <= end)
    range_where = and_(*range_conditions) if range_conditions else None

    def _apply(stmt):
        return stmt.where(range_where) if range_where is not None else stmt

    # 总活动数
    total_stmt = _apply(select(func.count(Activity.id)))
    total_activities = (await db.execute(total_stmt)).scalar_one() or 0

    # 今日活动数（按 report_date == today，与筛选窗口无关）
    today_stmt = select(func.count(Activity.id)).where(Activity.report_date == today)
    today_activities = (await db.execute(today_stmt)).scalar_one() or 0

    # 活跃人员数
    active_staff_stmt = _apply(select(func.count(distinct(Activity.staff_id))))
    active_staff_count = (await db.execute(active_staff_stmt)).scalar_one() or 0

    # 不重复商机数（opportunity_id 优先，回退 opportunity 名称）
    opp_stmt = _apply(
        select(func.count(distinct(func.coalesce(Activity.opportunity_id, Activity.opportunity))))
        .where(
            or_(
                Activity.opportunity_id.isnot(None),
                Activity.opportunity.isnot(None),
            )
        )
    )
    opportunity_count = (await db.execute(opp_stmt)).scalar_one() or 0

    # 按行为类型
    by_type_stmt = _apply(
        select(Activity.activity_type, func.count(Activity.id))
        .group_by(Activity.activity_type)
    )
    by_type_rows = (await db.execute(by_type_stmt)).all()
    by_type = {row[0]: int(row[1]) for row in by_type_rows if row[0]}

    # 按部门
    by_dept_stmt = (
        select(Staff.department, func.count(Activity.id))
        .select_from(Activity)
        .join(Staff, Staff.id == Activity.staff_id)
        .group_by(Staff.department)
    )
    if range_where is not None:
        by_dept_stmt = by_dept_stmt.where(range_where)
    by_dept_rows = (await db.execute(by_dept_stmt)).all()
    by_department = {row[0]: int(row[1]) for row in by_dept_rows if row[0]}

    # 按人员（TOP 10）
    by_staff_stmt = (
        select(Staff.name, func.count(Activity.id).label("cnt"))
        .select_from(Activity)
        .join(Staff, Staff.id == Activity.staff_id)
        .group_by(Staff.name)
        .order_by(func.count(Activity.id).desc())
        .limit(10)
    )
    if range_where is not None:
        by_staff_stmt = by_staff_stmt.where(range_where)
    by_staff_rows = (await db.execute(by_staff_stmt)).all()
    by_staff = [{"name": row[0], "count": int(row[1])} for row in by_staff_rows]

    # 近 7 天趋势（含今天，从 6 天前 → 今天）
    seven_days_ago = today - timedelta(days=6)
    trend_stmt = (
        select(Activity.report_date, func.count(Activity.id))
        .where(Activity.report_date >= seven_days_ago)
        .where(Activity.report_date <= today)
        .group_by(Activity.report_date)
    )
    trend_rows = (await db.execute(trend_stmt)).all()
    trend_map = {row[0]: int(row[1]) for row in trend_rows if row[0] is not None}
    daily_trend: list[dict[str, Any]] = []
    for offset in range(7):
        d = seven_days_ago + timedelta(days=offset)
        daily_trend.append({"date": d.isoformat(), "count": trend_map.get(d, 0)})

    recent_stmt = (
        select(Activity, Staff)
        .join(Staff, Staff.id == Activity.staff_id, isouter=True)
        .order_by(Activity.report_date.desc(), Activity.id.desc())
        .limit(10)
    )
    if range_where is not None:
        recent_stmt = recent_stmt.where(range_where)
    recent_rows = (await db.execute(recent_stmt)).all()
    recent_activities = [_activity_to_dict(activity, staff) for activity, staff in recent_rows]

    return {
        "total_activities": int(total_activities),
        "today_activities": int(today_activities),
        "active_staff_count": int(active_staff_count),
        "total_users": int(active_staff_count),  # 前端兼容别名
        "opportunity_count": int(opportunity_count),
        "total_opportunities": int(opportunity_count),  # 前端兼容别名
        "by_type": by_type,
        "action_type_breakdown": by_type,  # 前端兼容别名
        "by_department": by_department,
        "department_breakdown": by_department,  # 前端兼容别名
        "by_staff": by_staff,
        "daily_trend": daily_trend,
        "recent_activities": recent_activities,
    }


# ---------------------------------------------------------------------------
# Dashboard 别名（前端调 /activities/dashboard）
# ---------------------------------------------------------------------------


@router.get("/dashboard")
async def dashboard(
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(require_permission("dashboard:view")),
) -> dict[str, Any]:
    """统计看板（/stats 的别名，便于前端对接）。"""

    return await get_stats(start_date=start_date, end_date=end_date, db=db)


# ---------------------------------------------------------------------------
# 行为类型字典
# ---------------------------------------------------------------------------


@router.get("/action-types")
async def get_action_types(
    _user: dict = Depends(require_permission("dashboard:view")),
) -> list[dict[str, str]]:
    """获取所有行为类型及其中文标签和颜色。"""

    return ACTION_TYPES


# ---------------------------------------------------------------------------
# Excel 导出
# ---------------------------------------------------------------------------


@router.get("/export")
async def export_excel(
    start_date: str | None = None,
    end_date: str | None = None,
    staff_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(require_permission("dashboard:export")),
) -> StreamingResponse:
    """导出筛选后的活动记录为 Excel（xlsx）。"""

    # 延迟导入：openpyxl 较重，避免影响应用冷启动。
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    start = _parse_date(start_date, "start_date")
    end = _parse_date(end_date, "end_date")

    conditions = []
    if staff_id is not None:
        conditions.append(Activity.staff_id == staff_id)
    if start is not None:
        conditions.append(Activity.report_date >= start)
    if end is not None:
        conditions.append(Activity.report_date <= end)

    stmt = (
        select(Activity, Staff)
        .join(Staff, Staff.id == Activity.staff_id, isouter=True)
        .order_by(Activity.report_date.desc(), Activity.id.desc())
    )
    if conditions:
        stmt = stmt.where(and_(*conditions))
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "活动记录"

    headers = ["日期", "人员", "部门", "行为类型", "客户", "商机", "描述"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for activity, staff in rows:
        ws.append(
            [
                activity.report_date.isoformat() if activity.report_date else "",
                staff.name if staff else "",
                staff.department if staff else "",
                activity.activity_type or "",
                activity.target or "",
                activity.opportunity or "",
                activity.description or "",
            ]
        )

    # 自适应列宽（粗略估算，避免逐行精确计算的额外成本）
    widths = [12, 14, 14, 14, 24, 24, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename_parts = ["activities"]
    if start:
        filename_parts.append(start.isoformat())
    if end:
        filename_parts.append(end.isoformat())
    if not start and not end:
        filename_parts.append(datetime.now().strftime("%Y%m%d"))
    filename = "_".join(filename_parts) + ".xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
